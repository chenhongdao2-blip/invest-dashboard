"""Extract an analyst Excel model into a uniform ModelView JSON.

Run:
    uv run --with openpyxl --with pyyaml python3 jobs/extract_model.py \
        --ticker VEEV --xlsx "~/Downloads/VEEV US Mar2026.xlsx"

Design (see docs/model-viz-design-2026-06-01.md):
  - openpyxl data_only=True reads CACHED formula values. Analyst models are
    almost all cross-sheet formulas; cached values exist ONLY if the file was
    opened+saved in Excel/LibreOffice. We FAIL LOUD if the None ratio is high
    rather than silently emit an empty JSON.
  - Per-model mapping lives in config/models/<TICKER>.yml (heterogeneous models
    => one config each; the period-axis discovery is generic).
  - Output -> data/models/<TICKER>.json  (gitignored; not for public Cloud, MVP).
"""
from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
import yaml

REPO = Path(__file__).resolve().parent.parent
NONE_FAIL_RATIO = 0.20


def _num(v):
    return round(float(v), 4) if isinstance(v, (int, float)) and not isinstance(v, bool) else None


def _cell(ws, ref):
    return ws[ref].value


def _range_vals(ws, ref):
    """Flat list of values for an A1 range like 'C15:I21' (row-major)."""
    return [[c.value for c in row] for row in ws[ref]]


def _fy_num(lab):
    m = re.search(r"FY(\d+)", str(lab))
    return int(m.group(1)) if m else None


def _q_key(lab):
    """Chronological sort key for a quarter label like '1Q26' -> 26*4 + (1-1)."""
    m = re.match(r"([1-4])Q(\d{2})", str(lab).strip())
    return None if not m else int(m.group(2)) * 4 + (int(m.group(1)) - 1)


def discover_quarters(ws, header_row, pattern, actual_through_q=None):
    """Discover quarterly period columns (e.g. '1Q26'). Quarters at/before
    `actual_through_q` (the last reported fiscal quarter) are ACTUAL, later are
    estimates. No 'E' suffix exists on quarter labels in this model, so the
    actual/forecast split is purely chronological."""
    rx = re.compile(pattern)
    at = _q_key(actual_through_q)
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and rx.match(v.strip()):
            lab = v.strip()
            k = _q_key(lab)
            is_actual = at is not None and k is not None and k <= at
            out.append({"label": lab, "col": c, "kind": "quarter",
                        "actual_est": "A" if is_actual else "E"})
    return out


def discover_periods(ws, header_row, pattern, actual_through=None):
    """Discover annual period columns. actual/forecast is normally inferred from the
    trailing "E", but `actual_through` (e.g. "FY26") overrides it: FYs at/before it
    are reclassified ACTUAL and the stray "E" is stripped from their display label
    (a model may still suffix a now-reported FY with "E" if it wasn't relabeled)."""
    rx = re.compile(pattern)
    at = _fy_num(actual_through)
    out = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if isinstance(v, str) and rx.match(v.strip()):
            lab = v.strip()
            yr = _fy_num(lab)
            is_actual = (at is not None and yr is not None and yr <= at) \
                or (at is None and not lab.endswith("E"))
            disp = lab[:-1] if (is_actual and lab.endswith("E")) else lab
            out.append({"label": disp, "col": c, "kind": "annual",
                        "actual_est": "A" if is_actual else "E"})
    return out


def _series(ws, row, periods):
    """{period_label: numeric_value} for a given row across the period columns."""
    return {p["label"]: _num(ws.cell(row, p["col"]).value) for p in periods}


def _count_none(d):
    return sum(1 for v in d.values() if v is None), len(d)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--ticker", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--config")
    ap.add_argument("--out")
    args = ap.parse_args()

    cfg_path = Path(args.config) if args.config else REPO / "config" / "models" / f"{args.ticker}.yml"
    xlsx_path = Path(os.path.expanduser(args.xlsx))
    out_path = Path(args.out) if args.out else REPO / "data" / "models" / f"{args.ticker}.json"

    if not cfg_path.exists():
        sys.exit(f"ERROR: config not found: {cfg_path}")
    if not xlsx_path.exists():
        sys.exit(f"ERROR: xlsx not found: {xlsx_path}")

    cfg = yaml.safe_load(cfg_path.read_text())
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    none_hits = none_total = 0

    def track(series):
        nonlocal none_hits, none_total
        h, t = _count_none(series)
        none_hits += h
        none_total += t
        return series

    # ── periods (annual axis) ──
    pc = cfg["periods"]
    hdr_ws = wb[pc["sheet"]]
    periods = discover_periods(hdr_ws, pc["header_row"], pc["pattern"],
                               pc.get("actual_through"))
    if not periods:
        sys.exit("ERROR: no periods discovered — check periods config / header_row")
    first_e = next((p["label"] for p in periods if p["actual_est"] == "E"), None)

    # ── periods_q (optional quarterly axis) — trimmed to the contiguous span where
    # Reported Revenue is non-null (drops empty pre-IPO quarters / blank far-future).
    periods_q = []
    if pc.get("quarter_pattern"):
        raw_q = discover_quarters(hdr_ws, pc["header_row"], pc["quarter_pattern"],
                                  pc.get("actual_through_q"))
        if raw_q:
            qrev = _series(wb[cfg["revenue"]["sheet"]], cfg["revenue"]["total_row"], raw_q)
            idx = [i for i, p in enumerate(raw_q) if qrev.get(p["label"]) is not None]
            if idx:
                periods_q = raw_q[idx[0]:idx[-1] + 1]
    first_e_q = next((p["label"] for p in periods_q if p["actual_est"] == "E"), None)

    def _qvals(ws, row):
        return _series(ws, row, periods_q) if periods_q else {}

    def full_series(ws, row):
        """Annual series (counted by the None-gate) merged with the quarterly
        series (NOT counted — quarterly sparsity is legitimate, not a fetch failure)."""
        v = track(_series(ws, row, periods))
        v.update(_qvals(ws, row))
        return v

    # ── (a) revenue ──
    rv = cfg["revenue"]
    rs = wb[rv["sheet"]]                                   # totals (clean curated)
    seg_ws = wb[rv.get("segment_sheet", rv["sheet"])]      # 2×2 segments (e.g. Driver)
    revenue_breakdown = []
    for seg in rv["segments"]:
        entry = {
            "key": seg.get("key"),
            "segment": seg.get("segment", seg["en"]),       # strategic group: Commercial / R&D
            "segment_cn": seg.get("seg_cn", seg["cn"]),
            "label_en": seg["en"], "label_cn": seg["cn"],
            "basis": seg["basis"],
            "values": full_series(seg_ws, seg["row"]),
        }
        if "yoy_row" in seg:
            entry["yoy"] = {**_series(seg_ws, seg["yoy_row"], periods),
                            **_qvals(seg_ws, seg["yoy_row"])}
        revenue_breakdown.append(entry)
    revenue_total = {
        "values": full_series(rs, rv["total_row"]),
        "yoy": {**_series(rs, rv["total_yoy_row"], periods),
                **_qvals(rs, rv["total_yoy_row"])},
    }

    # ── (b) margins ──
    ms = wb[cfg["margins"]["sheet"]]
    margins = []
    for m in cfg["margins"].get("gaap", []):
        margins.append({"metric": m["key"], "metric_cn": m["cn"], "metric_en": m["en"],
                        "basis": "gaap", "values": full_series(ms, m["row"])})
    rev_for_margin = revenue_total["values"]
    for m in cfg["margins"].get("non_gaap_computed", []):
        num = {**_series(ms, m["num_row"], periods), **_qvals(ms, m["num_row"])}
        vals = {}
        for p in periods + periods_q:               # annual + quarterly columns
            n, d = num.get(p["label"]), rev_for_margin.get(p["label"])
            vals[p["label"]] = round(n / d, 4) if (n is not None and d) else None
        margins.append({"metric": m["key"], "metric_cn": m["cn"], "metric_en": m["en"],
                        "basis": "non_gaap", "values": vals})

    # ── (c) forecast ──
    fs = wb[cfg["forecast"]["sheet"]]
    forecast = []
    for ln in cfg["forecast"]["lines"]:
        forecast.append({"line": ln["key"], "line_cn": ln["cn"], "line_en": ln["en"],
                         "basis": ln["basis"], "unit": ln.get("unit", "millions"),
                         "values": full_series(fs, ln["row"])})

    # ── (d) analyst ratios + derived ─────────────────────────────────────────
    # Raw rows via full_series (annual tracked by the None-gate + quarterly merged,
    # quarterly NOT counted). Derived computed over the ANNUAL `periods` only, so
    # their quarterly keys stay absent and the page greys them out naturally.
    ratios = None
    if "ratios" in cfg:
        rc = cfg["ratios"]
        rws = wb[rc["sheet"]]
        ann = [p["label"] for p in periods]                 # annual axis labels

        # raw metric value dicts, keyed by metric key — reused by the derived calcs.
        raw_vals: dict[str, dict] = {}
        groups = []
        for g in rc.get("groups", []):
            metrics = []
            for mt in g["metrics"]:
                vals = full_series(rws, mt["row"])           # annual + quarterly
                raw_vals[mt["key"]] = vals
                metrics.append({"key": mt["key"], "cn": mt["cn"], "en": mt["en"],
                                "fmt": mt["fmt"], "freq": mt["freq"], "values": vals})
            groups.append({"key": g["key"], "cn": g["cn"], "en": g["en"],
                           "metrics": metrics})

        # input-only rows (feed the derived calcs; NOT emitted as display metrics).
        # Annual-only — tracked by the None-gate like any other mapped annual cell.
        inputs = {name: track(_series(rws, row, periods))
                  for name, row in rc.get("inputs", {}).items()}
        # total-revenue YoY (already curated as revenue_total['yoy']) feeds rule40.
        rev_yoy = revenue_total["yoy"]

        def _safe_div(n, d):
            return round(n / d, 4) if (n is not None and d not in (None, 0)) else None

        def _feed(name):
            """Resolve a derived feeder name → its annual period→value dict."""
            if name == "__rev_yoy__":
                return rev_yoy
            return raw_vals.get(name) or inputs.get(name) or {}

        gmap = {g["key"]: g for g in groups}
        for d in rc.get("derived", []):
            typ = d["type"]
            vals = {}
            if typ == "ratio":
                num, den = _feed(d["num"]), _feed(d["den"])
                for lbl in ann:
                    vals[lbl] = _safe_div(num.get(lbl), den.get(lbl))
            elif typ == "diff":
                a, b = _feed(d["a"]), _feed(d["b"])
                # freq=q derived (e.g. sbc_wedge): compute over annual + quarterly.
                span = ann if d.get("freq") != "q" else \
                    sorted(set(a) | set(b))
                for lbl in span:
                    av, bv = a.get(lbl), b.get(lbl)
                    vals[lbl] = round(av - bv, 4) if (av is not None and bv is not None) else None
            elif typ == "yoy":
                src = _feed(d["src"])
                for i, lbl in enumerate(ann):
                    prev = ann[i - 1] if i > 0 else None
                    cur, pv = src.get(lbl), (src.get(prev) if prev else None)
                    vals[lbl] = round(cur / pv - 1, 4) if (cur is not None and pv not in (None, 0)) else None
            elif typ == "rule40":
                growth, margin = _feed(d["growth_src"]), _feed(d["margin_src"])
                for lbl in ann:
                    gv, mv = growth.get(lbl), margin.get(lbl)
                    vals[lbl] = round(gv + mv, 4) if (gv is not None and mv is not None) else None
            entry = {"key": d["key"], "cn": d["cn"], "en": d["en"], "fmt": d["fmt"],
                     "freq": d.get("freq", "a"), "values": vals}
            gmap[d["group"]]["metrics"].append(entry)

        ratios = {"groups": groups}

    # ── GAAP → non-GAAP bridge ──
    gaap_bridge = None
    if "gaap_bridge" in cfg:
        bc = cfg["gaap_bridge"]
        bs = wb[bc["sheet"]]
        period_lbl = _cell(bs, bc["period_label_cell"])
        items = []
        for it in bc["rows"]:
            items.append({"item": it["item"], "item_en": it["en"],
                          "value": _num(bs[f"{bc['value_col']}{it['row']}"].value),
                          "total": it.get("total", False)})
        gaap_bridge = {"period": str(period_lbl), "items": items}

    # ── DCF + sensitivity + EV breakdown ──
    dcf = None
    if "dcf" in cfg:
        dc = cfg["dcf"]
        ds = wb[dc["sheet"]]
        sens_cfg = dc["sensitivity"]
        x_vals = [_num(v) for row in _range_vals(ds, sens_cfg["x_cells"]) for v in row]
        y_vals = [_num(v) for row in _range_vals(ds, sens_cfg["y_cells"]) for v in row]
        grid = [[_num(v) for v in row] for row in _range_vals(ds, sens_cfg["grid_cells"])]
        ev_break = []
        evb = dc.get("ev_breakdown")
        if evb:
            r0, r1 = (int(x) for x in str(evb["rows"]).split(":"))
            for r in range(r0, r1 + 1):
                seg = _cell(ds, f"{evb['seg_col']}{r}")
                if seg:
                    ev_break.append({"segment": str(seg),
                                     "npv": _num(ds[f"{evb['npv_col']}{r}"].value),
                                     "pct": _num(ds[f"{evb['pct_col']}{r}"].value)})
        dcf = {
            "wacc": _num(_cell(ds, dc["wacc_cell"])),
            "terminal_growth": _num(_cell(ds, dc["terminal_growth_cell"])),
            "target_price": _num(_cell(ds, dc["target_price_cell"])),
            "current_price": _num(_cell(ds, dc["current_price_cell"])),
            "ev": _num(_cell(ds, dc["ev_cell"])),
            "sensitivity": {"axis_x": sens_cfg["axis_x"], "axis_y": sens_cfg["axis_y"],
                            "x": x_vals, "y": y_vals, "grid": grid,
                            "base_x": sens_cfg.get("base_x"), "base_y": sens_cfg.get("base_y")},
            "ev_breakdown": ev_break,
            "scenario_note_cn": (str(_cell(ds, dc["scenario_note_cell"]) or "").strip()
                                 if dc.get("scenario_note_cell") else None),
        }

    # ── None-ratio gate ──
    ratio = (none_hits / none_total) if none_total else 1.0
    coverage = "full" if ratio < NONE_FAIL_RATIO else "low"
    if ratio >= NONE_FAIL_RATIO:
        sys.exit(f"ERROR: {ratio:.0%} of mapped cells are None ({none_hits}/{none_total}). "
                 "The xlsx likely has no cached formula values — open it in Excel/LibreOffice "
                 "and SAVE once, then re-run. (Aborting; not emitting partial JSON.)")

    meta = dict(cfg["meta"])
    meta.update({"ticker": args.ticker, "first_estimate_period": first_e,
                 "first_estimate_quarter": first_e_q,
                 "coverage": coverage, "confidence": "medium",
                 "none_ratio": round(ratio, 3)})

    def _ax(ps):
        return [{"label": p["label"], "kind": p["kind"], "actual_est": p["actual_est"]}
                for p in ps]

    doc = {"meta": meta, "periods": _ax(periods), "periods_q": _ax(periods_q),
           "revenue_total": revenue_total, "revenue_breakdown": revenue_breakdown,
           "margins": margins, "forecast": forecast, "ratios": ratios,
           "gaap_bridge": gaap_bridge, "dcf": dcf}

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(doc, ensure_ascii=False, indent=2))
    print(f"OK -> {out_path}  ({none_total} cells, {ratio:.1%} None, coverage={coverage})")
    print(f"   periods: {len(periods)} ({periods[0]['label']}..{periods[-1]['label']}), first E = {first_e}")
    if periods_q:
        print(f"   quarters: {len(periods_q)} ({periods_q[0]['label']}..{periods_q[-1]['label']}), first E = {first_e_q}")
    print(f"   revenue segs: {len(revenue_breakdown)}, margins: {len(margins)}, "
          f"forecast lines: {len(forecast)}, dcf: {'yes' if dcf else 'no'} (TP={dcf['target_price'] if dcf else '-'})")
    if ratios:
        _ng = len(ratios["groups"])
        _nm = sum(len(g["metrics"]) for g in ratios["groups"])
        print(f"   ratios: {_ng} groups, {_nm} metrics (incl. derived)")


if __name__ == "__main__":
    main()
