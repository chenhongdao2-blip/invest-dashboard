"""2025 中国创新药企员工人数变化（FY2024 vs FY2025）— 扩招 vs 收缩 发散条形图 + Excel。

仿照外资 big pharma "2025 Headcount Changes: Hirers and Cutters" 发散条形图，
做一张中国 12 家创新药/biotech 的对应版。

数据纪律（research-data.md）：每家 FY2024 / FY2025 员工总数均取自 HIGH 源，
逐行标注 source；无任何编造 / 框架默认值。口径 = 集团合并在职员工总数（年报披露）。

口径校准（关键）：iFind 日频「员工总数」在年末日期回填当年年报值（已用翰森年报
原文 9,347 校验），故年末取值 = 当年 FY 数；纯 H biotech 中 iFind NL 解析故障的，
直接取年报业绩公告 / ESG 报告原文。

来源标签:
  报告期   = iFind 财务「员工总数」报告期口径(20241231/20251231 两行)
  年报      = 港交所年度业绩公告 / 年报原文「雇员及薪酬」段
  ESG      = 公司 ESG 报告「员工总数」精确披露
  日频      = iFind 日频「员工总数」年末回填值

Run:
    uv run --with pandas --with openpyxl --with matplotlib \
        python jobs/cn_pharma_headcount_2025.py
"""

from __future__ import annotations

from pathlib import Path

REPO = Path(__file__).resolve().parent.parent
OUT_PNG = REPO / "data" / "external" / "cn_pharma_headcount_2025.png"
OUT_XLSX = REPO / "data" / "external" / "cn_pharma_headcount_2025.xlsx"
OUT_CSV = REPO / "data" / "external" / "cn_pharma_headcount.csv"   # committed; read by the Streamlit app
ASOF = "2026-06-03"

# 中文名 → 英文名（看板 i18n EN 标签用）
EN_NAME = {
    "信达生物": "Innovent", "百济神州": "BeiGene", "康方生物": "Akeso",
    "恒瑞医药": "Hengrui", "翰森制药": "Hansoh", "复宏汉霖": "Henlius",
    "康诺亚": "Keymed", "科伦博泰": "Kelun-Biotech", "诺诚健华": "InnoCare",
    "荣昌生物": "RemeGen", "石药集团": "CSPC", "中国生物制药": "Sino Biopharm",
}

# (中文名, 代码, FY2024, FY2025, 来源)  —— 全部 HIGH 源，逐行可溯
ROWS = [
    ("信达生物",   "01801.HK", 5659,  7502,  "年报：2025年报「7,502名(2024:5,659)雇员」"),
    ("百济神州",   "688235.SH", 11075, 11825, "报告期：iFind 员工总数 20241231/20251231"),
    ("康方生物",   "09926.HK", 3035,  3761,  "年报：2025业绩公告「3,761名雇员」(2024:3,035)"),
    ("恒瑞医药",   "600276.SH", 20238, 20602, "报告期：iFind 员工总数 20241231/20251231"),
    ("翰森制药",   "03692.HK", 8989,  9347,  "年报：2025业绩公告「九千三百四十七名全职雇员」"),
    ("复宏汉霖",   "02696.HK", 3515,  3762,  "报告期：iFind 员工总数 20241231/20251231"),
    ("康诺亚",     "02162.HK", 1258,  1469,  "报告期：iFind 员工总数 20241231/20251231"),
    ("科伦博泰",   "06990.HK", 1837,  2045,  "日频：iFind 员工总数 年末回填(1,837→2,045)"),
    ("诺诚健华",   "688428.SH", 1146,  1259,  "报告期：iFind 员工总数 20241231/20251231"),
    ("荣昌生物",   "688331.SH", 2999,  3048,  "报告期：iFind 员工总数 20241231/20251231"),
    ("石药集团",   "01093.HK", 21379, 19693, "ESG：2025 ESG报告 员工总数 19,693(2024:21,379)"),
    ("中国生物制药", "01177.HK", 24379, 21435, "年报：2025业绩公告「21,435名雇员」(2024:24,379)"),
]

ANCHOR_NOTE = (
    "口径：集团合并在职员工总数，FY2024 vs FY2025（年末）。来源：港交所年度业绩公告/年报"
    "、公司 ESG 报告、iFind。截至 2026-06-03。"
)

GREEN = "#2E9E5B"   # 扩招
RED = "#D8473E"     # 收缩


def render_png(data):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import rcParams

    rcParams["font.sans-serif"] = ["Arial Unicode MS"]
    rcParams["axes.unicode_minus"] = False

    # sort ascending so largest hirer ends on TOP (barh draws bottom-up)
    d = sorted(data, key=lambda r: r[4])  # r[4] = delta
    labels = [r[0] for r in d]
    deltas = [r[4] for r in d]
    colors = [GREEN if v > 0 else (RED if v < 0 else "#9AA0A6") for v in deltas]

    fig, ax = plt.subplots(figsize=(9.6, 6.4), dpi=200)
    fig.patch.set_facecolor("white")
    ax.set_facecolor("white")

    y = range(len(labels))
    ax.barh(list(y), deltas, color=colors, height=0.66, zorder=3)
    ax.set_yticks(list(y))
    ax.set_yticklabels(labels, fontsize=12)

    ax.axvline(0, color="#333333", lw=1.0, zorder=4)

    span = max(abs(min(deltas)), abs(max(deltas)))
    pad = span * 0.16
    ax.set_xlim(min(deltas) - pad * 1.7, max(deltas) + pad)

    # value labels at bar ends
    for i, v in enumerate(deltas):
        ha = "left" if v >= 0 else "right"
        off = span * 0.012
        ax.text(v + (off if v >= 0 else -off), i, f"{v:+,}",
                va="center", ha=ha, fontsize=11, fontweight="bold",
                color="#1A1A1A")

    ax.set_title("2025 中国创新药企员工人数变化：扩招 vs 收缩",
                 fontsize=15.5, fontweight="bold", pad=16, color="#1A1A1A")
    ax.set_xlabel("员工人数变化（人，FY2024 → FY2025）", fontsize=12)

    ax.spines[["top", "right", "left"]].set_visible(False)
    ax.tick_params(length=0)
    ax.xaxis.grid(True, color="#E6E6E6", lw=0.8, zorder=0)
    ax.set_axisbelow(True)

    fmt = lambda x, _: f"{int(x):,}"
    ax.xaxis.set_major_formatter(plt.FuncFormatter(fmt))

    fig.text(0.01, 0.015, ANCHOR_NOTE, fontsize=7.5, color="#8A8A8A")
    fig.subplots_adjust(left=0.16, right=0.965, top=0.91, bottom=0.115)
    fig.savefig(OUT_PNG, facecolor="white")
    plt.close(fig)


def write_xlsx(data):
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

    d = sorted(data, key=lambda r: -r[4])  # desc: hirers first
    wb = Workbook()
    ws = wb.active
    ws.title = "员工人数变化2025"

    head = ["公司", "代码", "FY2024 员工总数", "FY2025 员工总数", "变化", "变化%", "数据来源"]
    ws.append(head)
    hfill = PatternFill("solid", fgColor="F0E9DE")
    bottom = Border(bottom=Side(style="thin", color="D8CFC0"))
    for c in range(1, len(head) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, size=10, color="1A1A1A")
        cell.fill = hfill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bottom

    for name, code, fy24, fy25, delta, src in [(r[0], r[1], r[2], r[3], r[4], r[5]) for r in d]:
        pct = (fy25 - fy24) / fy24
        ws.append([name, code, fy24, fy25, delta, pct, src])

    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(r, 3).number_format = "#,##0"
        ws.cell(r, 4).number_format = "#,##0"
        ws.cell(r, 5).number_format = "+#,##0;-#,##0"
        ws.cell(r, 6).number_format = "+0.0%;-0.0%"
        delta = ws.cell(r, 5).value
        col = "2E9E5B" if delta > 0 else ("D8473E" if delta < 0 else "9AA0A6")
        ws.cell(r, 5).font = Font(bold=True, color=col)
    widths = [14, 11, 16, 16, 11, 9, 52]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # native bar chart of 变化 (col 5), categories = company (col 1)
    chart = BarChart()
    chart.type = "bar"
    chart.title = "2025 中国创新药企员工人数变化：扩招 vs 收缩 (FY2024→FY2025)"
    chart.y_axis.title = "员工人数变化(人)"
    chart.x_axis.delete = False
    chart.y_axis.delete = False
    chart.height = 11
    chart.width = 22
    chart.legend = None
    data_ref = Reference(ws, min_col=5, min_row=1, max_row=last)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, "I2")

    # source note row
    ws.cell(last + 2, 1, ANCHOR_NOTE).font = Font(size=9, color="8A8A8A")

    wb.save(OUT_XLSX)


def write_csv(data):
    """Committed tidy CSV the Streamlit app reads (cloud can't fetch iFind/年报 live)."""
    import csv

    d = sorted(data, key=lambda r: -r[4])  # desc: hirers first
    with OUT_CSV.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["name_cn", "name_en", "ticker", "fy2024", "fy2025", "delta", "pct", "source", "asof"])
        for n, c, a, b, dl, s in d:
            w.writerow([n, EN_NAME.get(n, n), c, a, b, dl, round((b - a) / a, 4), s, ASOF])


def main():
    data = [(n, c, a, b, b - a, s) for (n, c, a, b, s) in ROWS]
    render_png(data)
    write_xlsx(data)
    write_csv(data)
    print(f"[ok] {OUT_PNG}")
    print(f"[ok] {OUT_XLSX}")
    print(f"[ok] {OUT_CSV}")
    print("\n排序(变化降序):")
    for n, c, a, b, dl, s in sorted(data, key=lambda r: -r[4]):
        print(f"  {n:<8} {c:<11} {a:>7,} → {b:>7,}  {dl:+6,}  ({(b-a)/a:+.1%})")
    tot = sum(r[4] for r in data)
    hire = sum(1 for r in data if r[4] > 0)
    cut = sum(1 for r in data if r[4] < 0)
    print(f"\n  净变化合计 {tot:+,} 人 | 扩招 {hire} 家 / 收缩 {cut} 家")


if __name__ == "__main__":
    main()
