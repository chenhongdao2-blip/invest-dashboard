"""Bake the Healthcare-page "institutional positioning + relative performance" data.

Two committed outputs the Streamlit app reads (cloud can't fetch these live):

1. data/external/hc_index_comparison.csv
   Rebased-comparison source series, fixed anchor = 2025-08-01 (Jonah's "去年8月"):
     panel "hk"     : HSHCI.HK / HSI.HK / HSTECH.HK (恒生医疗 vs 恒生 vs 恒科) — iFind
     panel "msci"   : KURE / MCHI                    (MSCI 中国医疗 vs MSCI 中国, ETF 代理) — yfinance
     panel "nbi"    : ^NBI / ^IXIC / XBI             (NBI 大盘生科 + XBI 等权生科 vs 纳指) — yfinance
     panel "sphc"   : ^SP500-35 / ^GSPC              (S&P 500 Health Care vs S&P 500) — yfinance
     panel "ai_bio" : ^NBI / XBI / ^SOX             (生物科技 NBI+XBI vs AI 硬件 ^SOX 半导体) — yfinance

   A series may belong to MORE THAN ONE panel (^NBI / XBI appear in both "nbi" and
   "ai_bio"): PANEL_SERIES is the membership source of truth and each (series, panel)
   pair is emitted as its own rows, so the long CSV stays panel-sliceable downstream.

   HK indices come from iFind (HIGH reliability) because HSHCI/HSTECH are NOT on
   Yahoo (404 / empty). US indices come from yfinance. The HK raw pull is kept
   for provenance at data/external/hk_index_raw_ifind_<date>.csv.

2. data/external/china_fund_hc_positioning.csv
   12 offshore China-equity funds' healthcare over/underweight vs their own
   benchmark, as of 31 Mar 2026, extracted from the audited xlsx
   "MXCN&HSI sector perf_for HC May26.xlsx" (14-turn audited sheet — we do NOT
   re-extract from the 12 PDFs; the sheet is the source of truth).

Run locally (proxy needed for yfinance in CN):
    HTTP_PROXY=http://127.0.0.1:7897 HTTPS_PROXY=http://127.0.0.1:7897 \
    uv run --with yfinance --with openpyxl --with pandas \
    python jobs/build_hc_overview_data.py
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

import pandas as pd

REPO = Path(__file__).resolve().parent.parent
OUT = REPO / "data" / "external"
ANCHOR = "2025-08-01"
# yfinance `end` is EXCLUSIVE — use tomorrow so the latest completed session is included.
END = (date.today() + timedelta(days=1)).isoformat()

# iFind raw HK pull (provenance file already saved by the session).
HK_RAW = OUT / "hk_index_raw_ifind_20260601.csv"

# Desktop xlsx (external, not committed) — audited fund-positioning source.
XLSX = Path(
    "/Users/gcc/Desktop/Desktop - GCC的MacBook Pro/🗃️ 数据&脚本/"
    "fund_factsheets/MXCN&HSI sector perf_for HC May26.xlsx"
)

# series_id -> (en, cn, source). Panel membership is NO LONGER here — a series can
# live in >1 panel (see PANEL_SERIES below), so identity/source is decoupled from layout.
SERIES_META = {
    "HSHCI.HK":  ("Hang Seng Healthcare", "恒生医疗保健", "iFind"),
    "HSI.HK":    ("Hang Seng Index",      "恒生指数",     "iFind"),
    "HSTECH.HK": ("Hang Seng TECH",       "恒生科技",     "iFind"),
    "^NBI":      ("Nasdaq Biotech (NBI)", "纳斯达克生物科技", "yfinance"),
    "^IXIC":     ("Nasdaq Composite",     "纳斯达克综合",   "yfinance"),
    # XBI = S&P Biotech ETF (EQUAL-WEIGHT) — the SMID-cap breadth read vs ^NBI's
    # market-cap-weighted large-cap dominance. ETF price is USD + total-return.
    "XBI":       ("S&P Biotech (XBI)",    "标普生物科技 (XBI)", "yfinance · ETF"),
    # ^SOX = PHLX Semiconductor — the US "AI hardware" anchor (ai-researcher reviewed,
    # see benchmarks.py). An index (not an ETF), same quote convention as ^NBI / ^IXIC.
    "^SOX":      ("PHLX Semiconductor (SOX)", "费城半导体 (SOX)", "yfinance"),
    "^SP500-35": ("S&P 500 Health Care",  "标普500医疗保健", "yfinance"),
    "^GSPC":     ("S&P 500",              "标普500",      "yfinance"),
    # MSCI 口径 — investable ETF proxies (the MSCI index levels themselves aren't free /
    # daily-available; iFind doesn't carry them either). KURE tracks MSCI China All Shares
    # Health Care 10/40; MCHI tracks MSCI China. ETF prices are USD + total-return (div).
    "KURE":      ("MSCI China Health Care (KURE)", "MSCI中国医疗保健(KURE)", "yfinance · ETF"),
    "MCHI":      ("MSCI China (MCHI)",             "MSCI中国(MCHI)",         "yfinance · ETF"),
}

# panel -> [series_ids]. THE membership source of truth. ^NBI / XBI appear twice
# (nbi + ai_bio) on purpose — each (series, panel) pair is emitted as its own rows.
PANEL_SERIES = {
    "hk":     ["HSHCI.HK", "HSI.HK", "HSTECH.HK"],
    "msci":   ["KURE", "MCHI"],
    "nbi":    ["^NBI", "^IXIC", "XBI"],
    "sphc":   ["^SP500-35", "^GSPC"],
    "ai_bio": ["^NBI", "XBI", "^SOX"],
}
# Unique Yahoo-sourced series across ALL panels — startswith so "yfinance · ETF" matches.
US_TICKERS = sorted({sid for sids in PANEL_SERIES.values()
                     for sid in sids if SERIES_META[sid][2].startswith("yfinance")})


def build_index_comparison() -> pd.DataFrame:
    """Long tidy frame: date, series_id, name_en, name_cn, panel, close, source.

    Fetch each unique series' close ONCE (HK from iFind provenance, US from yfinance),
    then explode by PANEL_SERIES membership so a series shared across panels (^NBI /
    XBI in nbi + ai_bio) is emitted as one row-set per panel.
    """
    closes: dict[str, pd.DataFrame] = {}   # series_id -> df(date, close)

    # --- HK from iFind provenance CSV (per-series close) ---
    hk = pd.read_csv(HK_RAW)
    hk.columns = ["series_id", "name_raw", "date", "close"]
    hk["date"] = pd.to_datetime(hk["date"], format="%Y%m%d")
    for sid, g in hk[hk["series_id"].isin(SERIES_META)].groupby("series_id"):
        closes[sid] = g[["date", "close"]].sort_values("date").reset_index(drop=True)

    # --- US from yfinance (per-series close) ---
    import yfinance as yf

    d = yf.download(US_TICKERS, start=ANCHOR, end=END, auto_adjust=True,
                    progress=False, threads=True, group_by="ticker")
    for t in US_TICKERS:
        ser = d[t]["Close"].dropna() if t in d.columns.get_level_values(0) else pd.Series(dtype=float)
        if ser.empty:
            raise RuntimeError(f"yfinance returned empty for {t}")
        closes[t] = pd.DataFrame({"date": ser.index, "close": ser.values})

    # --- explode by panel membership (a series can land in >1 panel) ---
    rows: list[pd.DataFrame] = []
    for panel, sids in PANEL_SERIES.items():
        for sid in sids:
            c = closes.get(sid)
            if c is None or c.empty:
                raise RuntimeError(f"missing close series for {sid} (panel {panel})")
            sub = c.copy()
            sub["series_id"] = sid
            sub["panel"] = panel
            rows.append(sub)

    df = pd.concat(rows, ignore_index=True)
    df = df[df["date"] >= pd.Timestamp(ANCHOR)].copy()
    df["name_en"] = df["series_id"].map(lambda s: SERIES_META[s][0])
    df["name_cn"] = df["series_id"].map(lambda s: SERIES_META[s][1])
    df["source"] = df["series_id"].map(lambda s: SERIES_META[s][2])
    df = df.sort_values(["panel", "series_id", "date"]).reset_index(drop=True)
    df["date"] = df["date"].dt.strftime("%Y-%m-%d")
    return df[["date", "series_id", "name_en", "name_cn", "panel", "close", "source"]]


def build_fund_positioning() -> pd.DataFrame:
    """12 funds × HC over/underweight, from the audited 'fund positioning Mar 31 2026' sheet."""
    import openpyxl

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["fund positioning Mar 31 2026"]
    cols = ["fund", "aum_2026", "aum_2025", "aum_yoy", "benchmark",
            "fund_hc_w", "bm_hc_w", "ow_uw_2026", "deviation_2026", "data_date",
            "ow_uw_2025", "deviation_2025", "change_dev"]
    recs = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0 or i > 12:        # row 0 header, rows 1-12 the funds
            continue
        rec = dict(zip(cols, [c for c in row[:13]]))
        if not rec["fund"]:
            continue
        # strip the stray 中文 suffix one fund's name carries in the sheet
        rec["fund"] = str(rec["fund"]).replace("（无healthcare）", "").strip()
        recs.append(rec)
    df = pd.DataFrame(recs)
    # numeric coercion (N/A strings -> NaN)
    for c in ["fund_hc_w", "bm_hc_w", "deviation_2026", "deviation_2025", "change_dev", "aum_yoy"]:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    df["data_date"] = df["data_date"].astype(str).str.replace(" 00:00:00", "", regex=False)
    return df


# Generic fallback for the FULL (local) note if the xlsx footnote can't be read.
# The detailed fund-naming note lives in the (local-only) xlsx footnote cell, not in
# this committed script — read_full_source_note() pulls it from there so no fund
# names are hardcoded into source that ships to the public Cloud repo.
SOURCE_NOTE_FULL_FALLBACK = (
    "Fund fact sheets as at 31 Mar 2026 (a few as at 30 Apr 2026 / Q1 2026). One fund's "
    "healthcare weight is not disclosed numerically. Deviation = Fund HC weight − Benchmark HC weight."
)


def read_full_source_note() -> str:
    """The names-bearing source footnote from the (local) xlsx column A — never hardcoded."""
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["fund positioning Mar 31 2026"]
    for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        cell = row[0]
        if cell and isinstance(cell, str) and "fact sheet" in cell.lower() and len(cell) > 80:
            return cell.strip()
    return SOURCE_NOTE_FULL_FALLBACK


# Public source note (NO fund names) — committed, shipped to Cloud.
SOURCE_NOTE_PUBLIC = (
    "Offshore China-equity fund fact sheets as at 31 Mar 2026 (a few as at 30 Apr 2026 / Q1 2026). "
    "One fund's healthcare weight is not disclosed numerically. Fund names withheld; labelled "
    "Fund 1–12. Deviation = Fund HC weight − Benchmark HC weight."
)


def anonymise(pos: pd.DataFrame) -> pd.DataFrame:
    """Replace real fund names with 'Fund 1'..'Fund N' by row order (preserves the
    long-standing sheet convention: Fund 1 = the first row, etc.). Only the `fund`
    column carries identity — AUM / benchmark / weights are non-identifying and kept."""
    out = pos.copy()
    out["fund"] = [f"Fund {i}" for i in range(1, len(out) + 1)]
    return out


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)

    idx = build_index_comparison()
    idx_path = OUT / "hc_index_comparison.csv"
    idx.to_csv(idx_path, index=False)
    print(f"[ok] {idx_path}  ({len(idx)} rows, {idx['series_id'].nunique()} series)")
    for panel, g in idx.groupby("panel"):
        spread = g.groupby("series_id")["date"].agg(["min", "max", "count"])
        print(f"   panel {panel}:")
        print(spread.to_string().replace("\n", "\n      "))

    pos = build_fund_positioning()
    # FULL (real names) — gitignored, local-only. App prefers this when present.
    full_path = OUT / "china_fund_hc_positioning_full.csv"
    pos.to_csv(full_path, index=False)
    (OUT / "china_fund_hc_positioning_source_full.txt").write_text(read_full_source_note(), encoding="utf-8")
    print(f"[ok] {full_path}  ({len(pos)} funds, REAL names — gitignored)")
    # PUBLIC (Fund 1–12) — committed, shipped to Cloud.
    pos_path = OUT / "china_fund_hc_positioning.csv"
    anonymise(pos).to_csv(pos_path, index=False)
    (OUT / "china_fund_hc_positioning_source.txt").write_text(SOURCE_NOTE_PUBLIC, encoding="utf-8")
    print(f"[ok] {pos_path}  ({len(pos)} funds, ANON Fund 1–12 — committed)")

    # quick verdict sanity print
    have = pos.dropna(subset=["deviation_2026"])
    ow = (have["ow_uw_2026"].str.contains("OW", na=False)).sum()
    uw = (have["ow_uw_2026"] == "UW").sum()
    neu = (have["ow_uw_2026"].str.contains("Neutral", na=False)).sum()
    print(f"   verdict: {ow} OW / {uw} UW / {neu} Neutral (of {len(have)} with data)")


if __name__ == "__main__":
    main()
