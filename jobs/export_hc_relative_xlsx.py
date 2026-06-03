"""Export the Healthcare-page "相对表现" three panels to a single Excel workbook.

Reads data/external/hc_index_comparison.csv (the committed source the Streamlit app
plots) and reproduces the app's EXACT rebase convention — see app/lib/charts.py:
inner-join each panel's series on common trading days, then
    rebased = wide / wide.iloc[0] * 100        (anchor = first common date)

Output (one workbook, one sheet per panel + a README sheet):
    每个面板 sheet: 日期 | <各序列 rebased> | 空列 | <各序列原始收盘价>
                    + 一张原生 Excel 折线图(rebased)，hero 序列红线。

Run:
    uv run --with pandas --with openpyxl \
        python jobs/export_hc_relative_xlsx.py
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO = Path(__file__).resolve().parent.parent
SRC = REPO / "data" / "external" / "hc_index_comparison.csv"
OUT = REPO / "data" / "external" / "hc_relative_performance.xlsx"

ANCHOR = "2025-08-01"

# panel -> (中文标题, hero series_id, [series order])
PANELS = {
    "hk":   ("恒生医疗保健 vs 恒生 vs 恒生科技", "HSHCI.HK",  ["HSHCI.HK", "HSI.HK", "HSTECH.HK"]),
    "nbi":  ("纳斯达克生物科技 (NBI) vs 纳斯达克综合", "^NBI",  ["^NBI", "^IXIC"]),
    "sphc": ("标普 500 医疗保健 vs 标普 500", "^SP500-35", ["^SP500-35", "^GSPC"]),
}

# colours mirror the dashboard: hero = red, peers = grey
RED = "C0392B"
GREY = "8A8A8A"

INK = "1A1A1A"
HEAD_FILL = PatternFill("solid", fgColor="F0E9DE")  # warm header like the app
THIN = Side(style="thin", color="D8CFC0")
BORDER = Border(bottom=THIN)


def load() -> pd.DataFrame:
    df = pd.read_csv(SRC, parse_dates=["date"])
    return df


def panel_wide(df: pd.DataFrame, series_order: list[str]) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """Return (raw_close_wide, rebased_wide, {series_id: name_cn}) for one panel.

    Reproduces app/lib/charts.py: inner-join on common dates, rebase to first row.
    """
    sub = df[df["series_id"].isin(series_order)]
    names = {sid: sub[sub["series_id"] == sid]["name_cn"].iloc[0] for sid in series_order}
    wide = (
        sub.pivot(index="date", columns="series_id", values="close")
        .reindex(columns=series_order)  # preserve hero-first order
        .dropna()                       # inner-join: common trading days only
        .sort_index()
    )
    rebased = wide.divide(wide.iloc[0]) * 100.0
    return wide, rebased, names


def write_panel(wb, panel: str) -> None:
    title_cn, hero, order = PANELS[panel]
    raw, reb, names = PANEL_DATA[panel]
    ws = wb.create_sheet(title_cn[:28])  # sheet-name length cap

    n_series = len(order)
    # --- header row ---
    headers = ["日期"]
    headers += [f"{names[s]} (rebased)" for s in order]
    headers += [""]  # spacer
    headers += [f"{names[s]} (收盘价)" for s in order]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color=INK, size=10)
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    # --- data rows ---
    idx = reb.index
    for i, d in enumerate(idx):
        row = [d.date()]
        row += [round(float(reb.iloc[i][s]), 2) for s in order]
        row += [""]
        row += [round(float(raw.iloc[i][s]), 2) for s in order]
        ws.append(row)

    # --- formatting ---
    last_row = ws.max_row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        for c in range(2, 2 + n_series):
            ws.cell(row=r, column=c).number_format = "0.00"
        for c in range(3 + n_series, 3 + 2 * n_series):
            ws.cell(row=r, column=c).number_format = "#,##0.00"
    ws.column_dimensions["A"].width = 12
    for c in range(2, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "B2"

    # --- native Excel line chart of the rebased block ---
    chart = LineChart()
    chart.title = title_cn
    chart.style = 2
    chart.height = 9
    chart.width = 22
    chart.y_axis.title = "rebased (起点=100)"
    chart.x_axis.title = None
    chart.x_axis.number_format = "yyyy-mm"
    chart.x_axis.majorTimeUnit = "months"

    data_ref = Reference(ws, min_col=2, max_col=1 + n_series, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    for k, ser in enumerate(chart.series):
        sid = order[k]
        is_hero = sid == hero
        colour = RED if is_hero else GREY
        ser.graphicalProperties.line = LineProperties(
            solidFill=colour, w=28000 if is_hero else 14000,
        )
        if not is_hero:
            ser.graphicalProperties.line.prstDash = "dash" if k == 1 else "sysDot"
        ser.smooth = False

    chart.x_axis.delete = False
    chart.y_axis.delete = False
    anchor_col = get_column_letter(3 + 2 * n_series + 1)
    ws.add_chart(chart, f"{anchor_col}2")


def write_readme(wb) -> None:
    ws = wb.create_sheet("说明", 0)
    last = max(reb.index.max() for _, reb, _ in PANEL_DATA.values())
    lines = [
        ("恒生/美股 医疗保健 相对表现 — 数据导出", True),
        ("", False),
        (f"锚定日 (起点=100): {ANCHOR}（去年 8 月）", False),
        (f"数据截至: {last.date()}", False),
        ("来源: 港股指数 iFind（HSHCI/HSTECH 不在 Yahoo）; 美股指数 yfinance。", False),
        ("", False),
        ("Rebase 口径（与看板图完全一致）:", True),
        ("  1) 每个面板内的各序列按【共同交易日】inner-join 对齐;", False),
        ("  2) rebased = 收盘价 ÷ 共同首日收盘价 × 100;", False),
        ("  3) 锚点 = 对齐后的第一个共同交易日, 非各自首日。", False),
        ("", False),
        ("每个面板 sheet 同时给出 rebased 序列与原始收盘价, 右侧为原生 Excel 折线图。", False),
        ("hero 序列(医疗/生物)= 红线; 对标指数 = 灰色虚线。", False),
        ("", False),
        ("面板:", True),
    ]
    for _, (t, hero, order) in PANELS.items():
        lines.append((f"  · {t}  [{', '.join(order)}]", False))
    for r, (txt, bold) in enumerate(lines, start=1):
        c = ws.cell(row=r, column=1, value=txt)
        c.font = Font(bold=bold, size=13 if (bold and r == 1) else 10, color=INK)
    ws.column_dimensions["A"].width = 80


def main() -> None:
    from openpyxl import Workbook

    global PANEL_DATA
    df = load()
    PANEL_DATA = {p: panel_wide(df, order) for p, (_, _, order) in PANELS.items()}

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    for panel in PANELS:
        write_panel(wb, panel)
    write_readme(wb)
    wb.save(OUT)

    print(f"[ok] {OUT}")
    for p, (t, hero, order) in PANELS.items():
        raw, rb, names = PANEL_DATA[p]
        last = rb.iloc[-1]
        spreads = ", ".join(
            f"{names[s]} {last[hero] - last[s]:+.1f}pp" for s in order if s != hero
        )
        print(f"   {t}: {len(rb)} 行, hero={names[hero]} 末值 {last[hero]:.1f} | 对标 {spreads}")


if __name__ == "__main__":
    main()
