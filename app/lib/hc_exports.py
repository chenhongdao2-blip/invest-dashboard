"""In-memory .xlsx builders for the Healthcare page download buttons.

Each `*_xlsx()` returns raw bytes for st.download_button. Cached on the source
file's mtime so an in-place re-bake (jobs/build_hc_overview_data.py or
jobs/cn_pharma_headcount_2025.py) invalidates the cached workbook.

Reuses the page's own loaders (lib.hc_overview) so:
  - relative performance reproduces the EXACT app rebase (wide/iloc[0]*100 on the
    panel's common trading days — same convention as lib/charts.py);
  - positioning respects full-vs-anon automatically (local → real fund names,
    Cloud → Fund 1–12), since load_fund_positioning() already picks the file.

Each workbook ships a data sheet + a native Excel chart so the download is
deck-ready, not just a number dump.
"""

from __future__ import annotations

from io import BytesIO

import pandas as pd
import streamlit as st

from lib import hc_overview as hco

# Colours mirror the app's LOCKED convention (teal = up/OW/hire, red = down/UW/cut).
TEAL = "0D7680"
RED = "CC0000"
CMSI_RED = "C8102E"
GREY = "8A8A8A"
HEAD_FILL = "F0E9DE"
INK = "1A1A1A"

ANCHOR = "2025-08-01"
_REL_PANELS = {
    "恒生医疗 vs 恒生 vs 恒生科技": ("HSHCI.HK", ["HSHCI.HK", "HSI.HK", "HSTECH.HK"]),
    "NBI vs 纳斯达克综合": ("^NBI", ["^NBI", "^IXIC"]),
    "标普500医疗 vs 标普500": ("^SP500-35", ["^SP500-35", "^GSPC"]),
}


def _style_header(ws, ncol: int) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    fill = PatternFill("solid", fgColor=HEAD_FILL)
    border = Border(bottom=Side(style="thin", color="D8CFC0"))
    for c in range(1, ncol + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, size=10, color=INK)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


# ----------------------------------------------------------------------------- #
# 1) 相对表现 — 3 panels, rebased(=100) + raw close + native line chart
# ----------------------------------------------------------------------------- #
def _build_relative(wb) -> None:
    from openpyxl.chart import LineChart, Reference
    from openpyxl.drawing.line import LineProperties
    from openpyxl.utils import get_column_letter

    df = hco.load_index_comparison()
    name_cn = dict(zip(df["series_id"], df["name_cn"]))

    for sheet_name, (hero, order) in _REL_PANELS.items():
        # A series may recur across panels (^NBI / XBI live in both `nbi` and
        # `ai_bio`); isin() alone would pull both copies and pivot would choke on
        # duplicate (date, series_id) pairs. The copies carry identical closes, so
        # dedup-keep-first is exact and panel-agnostic.
        sub = (df[df["series_id"].isin(order)]
               .drop_duplicates(subset=["date", "series_id"], keep="first"))
        wide = (sub.pivot(index="date", columns="series_id", values="close")
                .reindex(columns=order).dropna().sort_index())
        if wide.empty:
            continue
        reb = wide.divide(wide.iloc[0]) * 100.0
        n = len(order)

        ws = wb.create_sheet(sheet_name[:28])
        headers = (["日期"] + [f"{name_cn[s]} (rebased)" for s in order]
                   + [""] + [f"{name_cn[s]} (收盘价)" for s in order])
        ws.append(headers)
        _style_header(ws, len(headers))
        for i, d in enumerate(reb.index):
            row = ([d.date()] + [round(float(reb.iloc[i][s]), 2) for s in order]
                   + [""] + [round(float(wide.iloc[i][s]), 2) for s in order])
            ws.append(row)

        last = ws.max_row
        for r in range(2, last + 1):
            ws.cell(r, 1).number_format = "yyyy-mm-dd"
            for c in range(2, 2 + n):
                ws.cell(r, c).number_format = "0.00"
            for c in range(3 + n, 3 + 2 * n):
                ws.cell(r, c).number_format = "#,##0.00"
        ws.column_dimensions["A"].width = 12
        for c in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 16
        ws.freeze_panes = "B2"

        chart = LineChart()
        chart.title = sheet_name
        chart.height, chart.width = 9, 22
        chart.y_axis.title = "rebased (起点=100)"
        chart.x_axis.number_format = "yyyy-mm"
        chart.x_axis.majorTimeUnit = "months"
        data_ref = Reference(ws, min_col=2, max_col=1 + n, min_row=1, max_row=last)
        cats = Reference(ws, min_col=1, min_row=2, max_row=last)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        for k, ser in enumerate(chart.series):
            is_hero = order[k] == hero
            ser.graphicalProperties.line = LineProperties(
                solidFill=(RED if is_hero else GREY), w=28000 if is_hero else 14000)
            if not is_hero:
                ser.graphicalProperties.line.prstDash = "dash" if k == 1 else "sysDot"
            ser.smooth = False
        chart.x_axis.delete = chart.y_axis.delete = False
        ws.add_chart(chart, f"{get_column_letter(3 + 2 * n + 1)}2")

    # --- HSHCI long-history sheet: 2021.7→now absolute level + rebased + line chart ---
    hist = hco.load_hshci_history()
    if not hist.empty:
        hd = hist.sort_values("date").reset_index(drop=True)
        wsh = wb.create_sheet("HSHCI 长周期 2021.7-今")
        wsh.append(["日期", "收盘点位", "rebased(起点=100)"])
        _style_header(wsh, 3)
        base = float(hd["close"].iloc[0])
        for _, r in hd.iterrows():
            wsh.append([r["date"].date(), float(r["close"]), round(float(r["close"]) / base * 100, 2)])
        lasth = wsh.max_row
        for rr in range(2, lasth + 1):
            wsh.cell(rr, 1).number_format = "yyyy-mm-dd"
            wsh.cell(rr, 2).number_format = "#,##0.00"
            wsh.cell(rr, 3).number_format = "0.00"
        wsh.column_dimensions["A"].width = 12
        wsh.column_dimensions["B"].width = 14
        wsh.column_dimensions["C"].width = 16
        wsh.freeze_panes = "A2"
        ch = LineChart()
        ch.title = "恒生医疗保健指数：2021.7 以来完整轨迹(绝对点位)"
        ch.height, ch.width = 9, 22
        ch.y_axis.title = "指数点位"
        ch.x_axis.number_format = "yyyy-mm"
        ch.x_axis.majorTimeUnit = "months"
        ch.add_data(Reference(wsh, min_col=2, min_row=1, max_row=lasth), titles_from_data=True)
        ch.set_categories(Reference(wsh, min_col=1, min_row=2, max_row=lasth))
        ch.series[0].graphicalProperties.line = LineProperties(solidFill=CMSI_RED, w=26000)
        ch.x_axis.delete = ch.y_axis.delete = False
        wsh.add_chart(ch, "E2")

    ws0 = wb.create_sheet("说明", 0)
    asof = df["date"].max().date().isoformat()
    for r, line in enumerate([
        "恒生/美股 医疗保健 相对表现 — 数据导出",
        f"锚定日 (起点=100): {ANCHOR}（去年 8 月）  ·  数据截至: {asof}",
        "来源: 港股指数 iFind（HSHCI/HSTECH 不在 Yahoo）; 美股指数 yfinance。",
        "Rebase 口径(与看板图一致): 面板内各序列取共同交易日 inner-join → ÷ 共同首日 × 100。",
        "每个面板 sheet 同时给出 rebased 序列与原始收盘价, 右侧为原生折线图(医疗=红, 对标=灰虚线)。",
    ], start=1):
        from openpyxl.styles import Font
        ws0.cell(r, 1, line).font = Font(bold=(r == 1), size=13 if r == 1 else 10, color=INK)
    ws0.column_dimensions["A"].width = 86


# ----------------------------------------------------------------------------- #
# 2) 机构持仓 — 12 funds OW/UW + native diverging bar
# ----------------------------------------------------------------------------- #
def _build_positioning(wb) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    pos = hco.load_fund_positioning()
    ws = wb.active
    ws.title = "机构持仓 OW-UW"

    cols = [("fund", "基金"), ("aum_2026", "规模(USD)"), ("benchmark", "基准"),
            ("fund_hc_w", "基金HC%"), ("bm_hc_w", "基准HC%"),
            ("deviation_2026", "偏离(pp)"), ("ow_uw_2026", "超/低配"),
            ("change_dev", "较去年Δ(pp)"), ("data_date", "数据日期")]
    ws.append([c[1] for c in cols])
    _style_header(ws, len(cols))
    for _, row in pos.iterrows():
        ws.append([
            row.get("fund"), row.get("aum_2026"), row.get("benchmark"),
            None if pd.isna(row.get("fund_hc_w")) else float(row.get("fund_hc_w")),
            None if pd.isna(row.get("bm_hc_w")) else float(row.get("bm_hc_w")),
            None if pd.isna(row.get("deviation_2026")) else float(row.get("deviation_2026")) * 100,
            row.get("ow_uw_2026"),
            None if pd.isna(row.get("change_dev")) else float(row.get("change_dev")) * 100,
            row.get("data_date"),
        ])
    last = ws.max_row
    for r in range(2, last + 1):
        ws.cell(r, 4).number_format = '0.0"%"'
        ws.cell(r, 5).number_format = '0.0"%"'
        ws.cell(r, 6).number_format = "+0.0;-0.0"
        ws.cell(r, 8).number_format = "+0.0;-0.0"
    for i, w in enumerate([22, 12, 16, 10, 10, 11, 12, 12, 12], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    chart = BarChart()
    chart.type = "bar"
    chart.title = "医疗板块偏离度 (基金权重 − 基准权重, pp)"
    chart.y_axis.title = "偏离 (pp)"
    chart.height, chart.width = 11, 20
    chart.legend = None
    chart.add_data(Reference(ws, min_col=6, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
    chart.x_axis.delete = chart.y_axis.delete = False
    ws.add_chart(chart, "K2")

    src = hco.positioning_source()
    ws.cell(last + 2, 1, "颜色/口径: 青=超配(OW) / 红=低配(UW); 偏离 = 基金HC权重 − 基准HC权重。").font = Font(size=9, color=GREY)
    if src:
        ws.cell(last + 3, 1, "来源: " + src).font = Font(size=9, color=GREY)


# ----------------------------------------------------------------------------- #
# 3) 员工人数变化 — 12 names FY2024→FY2025 + native bar
# ----------------------------------------------------------------------------- #
def _build_headcount(wb) -> None:
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font

    hc = hco.load_headcount()
    ws = wb.active
    ws.title = "员工人数变化2025"

    head = ["公司", "英文名", "代码", "FY2024", "FY2025", "变化", "变化%", "数据来源"]
    ws.append(head)
    _style_header(ws, len(head))
    for _, r in hc.iterrows():
        ws.append([r["name_cn"], r["name_en"], r["ticker"], int(r["fy2024"]),
                   int(r["fy2025"]), int(r["delta"]), float(r["pct"]), r["source"]])
    last = ws.max_row
    for rr in range(2, last + 1):
        ws.cell(rr, 4).number_format = "#,##0"
        ws.cell(rr, 5).number_format = "#,##0"
        ws.cell(rr, 6).number_format = "+#,##0;-#,##0"
        ws.cell(rr, 7).number_format = "+0.0%;-0.0%"
        dl = ws.cell(rr, 6).value
        ws.cell(rr, 6).font = Font(bold=True, color=(TEAL if dl > 0 else RED if dl < 0 else GREY))
    for i, w in enumerate([14, 14, 11, 11, 11, 10, 9, 50], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    chart = BarChart()
    chart.type = "bar"
    chart.title = "员工人数变化 (FY2024→FY2025)"
    chart.y_axis.title = "变化(人)"
    chart.height, chart.width = 11, 20
    chart.legend = None
    chart.add_data(Reference(ws, min_col=6, min_row=1, max_row=last), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
    chart.x_axis.delete = chart.y_axis.delete = False
    ws.add_chart(chart, "J2")

    ws.cell(last + 2, 1, "口径: 集团合并在职员工总数(FY年末)。来源见每行「数据来源」列。青=扩招/红=收缩。").font = Font(size=9, color=GREY)


# ----------------------------------------------------------------------------- #
# 4) 日本医药 — 40 支明细 + 专栏指数 vs TOPIX/日经 (rebased) + native line chart
# ----------------------------------------------------------------------------- #
_JP_SUB_CN = {"pharma": "制药", "medtech": "医疗器械",
              "diagnostics": "诊断·检测", "distribution": "流通·服务"}


def _build_japan(wb) -> None:
    from openpyxl.chart import LineChart, Reference
    from openpyxl.drawing.line import LineProperties
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    from lib import db

    uni = hco.jp_universe()
    closes = db.get_close_series_usd(tuple(uni["ticker"]))
    rets = db.compute_returns(closes)

    # --- detail sheet: 40 names, USD 口径 ---
    ws = wb.create_sheet("日本医药 明细")
    head = ["代码", "名称", "Name", "子板块", "Last (USD)", "1日%", "5日%", "1月%", "年初至今%"]
    ws.append(head)
    _style_header(ws, len(head))
    m = uni.merge(rets, left_on="ticker", right_index=True, how="left")
    for _, r in m.iterrows():
        ws.append([
            r["ticker"], r["name_cn"], r["name_en"], _JP_SUB_CN.get(r["subsector"], r["subsector"]),
            None if pd.isna(r.get("last")) else round(float(r["last"]), 2),
            *[None if pd.isna(r.get(k)) else round(float(r[k]), 2)
              for k in ("1d_%", "5d_%", "1m_%", "ytd_%")],
        ])
    last = ws.max_row
    for rr in range(2, last + 1):
        ws.cell(rr, 5).number_format = "#,##0.00"
        for cc in range(6, 10):
            ws.cell(rr, cc).number_format = '+0.00"%";-0.00"%"'
    for i, w in enumerate([10, 16, 26, 12, 12, 9, 9, 9, 11], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    # --- composite vs TOPIX vs 日经 (all USD, rebased at common anchor) ---
    comp = hco.jp_composite(
        closes, weights=uni.set_index("ticker")["mcap_bn_jpy"]
        if "mcap_bn_jpy" in uni.columns else None)
    bench = hco.jp_benchmarks_usd()
    if comp is not None and bench:
        cols = {"日本医药专栏指数(40支市值加权)": comp}
        if "1305.T" in bench:
            cols["TOPIX (1305.T ETF代理)"] = bench["1305.T"]
        if "^N225" in bench:
            cols["日经225"] = bench["^N225"]
        wide = pd.DataFrame(cols).dropna().sort_index()
        if not wide.empty:
            reb = wide.divide(wide.iloc[0]) * 100.0
            ws2 = wb.create_sheet("专栏指数 vs TOPIX vs 日经")
            ws2.append(["日期"] + [f"{c} (rebased)" for c in reb.columns])
            _style_header(ws2, 1 + len(reb.columns))
            for i, d in enumerate(reb.index):
                ws2.append([d.date()] + [round(float(v), 2) for v in reb.iloc[i]])
            last2 = ws2.max_row
            for rr in range(2, last2 + 1):
                ws2.cell(rr, 1).number_format = "yyyy-mm-dd"
                for cc in range(2, 2 + len(reb.columns)):
                    ws2.cell(rr, cc).number_format = "0.00"
            ws2.column_dimensions["A"].width = 12
            for cc in range(2, 2 + len(reb.columns)):
                ws2.column_dimensions[get_column_letter(cc)].width = 26
            ws2.freeze_panes = "B2"
            chart = LineChart()
            chart.title = "日本医药专栏指数 vs TOPIX vs 日经225 (USD, rebased)"
            chart.height, chart.width = 9, 22
            chart.y_axis.title = "rebased (起点=100)"
            chart.x_axis.number_format = "yyyy-mm"
            chart.x_axis.majorTimeUnit = "months"
            chart.add_data(Reference(ws2, min_col=2, max_col=1 + len(reb.columns),
                                     min_row=1, max_row=last2), titles_from_data=True)
            chart.set_categories(Reference(ws2, min_col=1, min_row=2, max_row=last2))
            for k, ser in enumerate(chart.series):
                is_hero = k == 0
                ser.graphicalProperties.line = LineProperties(
                    solidFill=(CMSI_RED if is_hero else GREY), w=28000 if is_hero else 14000)
                if not is_hero:
                    ser.graphicalProperties.line.prstDash = "dash" if k == 1 else "sysDot"
                ser.smooth = False
            chart.x_axis.delete = chart.y_axis.delete = False
            ws2.add_chart(chart, f"{get_column_letter(3 + len(reb.columns))}2")

    asof = closes.index.max().date().isoformat() if not closes.empty else "—"
    ws.cell(last + 2, 1,
            f"口径: Last/收益均为 USD（JPY 经 USDJPY 换算, M1 口径）· 专栏指数=40支市值加权"
            f"(权重=2026/05市值快照) · "
            f"TOPIX 用 1305.T ETF 代理 · 截至 {asof}。").font = Font(size=9, color=GREY)
    ws.cell(last + 3, 1,
            "Universe: iFind 自选清单 2026/05 (42支) 剔除已退市 HOGY MEDICAL(3593)/久光制药(4530)。"
            ).font = Font(size=9, color=GREY)


def _save(wb) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(ttl=3600)
def relative_xlsx(mtime: float) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    _build_relative(wb)
    return _save(wb)


@st.cache_data(ttl=3600)
def positioning_xlsx(mtime: float) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    _build_positioning(wb)
    return _save(wb)


@st.cache_data(ttl=3600)
def headcount_xlsx(mtime: float) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    _build_headcount(wb)
    return _save(wb)


@st.cache_data(ttl=3600)
def japan_xlsx(mtime: float) -> bytes:
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    _build_japan(wb)
    return _save(wb)


def _mtime(*paths) -> float:
    """Max mtime across candidate paths (0.0 if none) — the cache key."""
    m = 0.0
    for p in paths:
        try:
            m = max(m, p.stat().st_mtime)
        except OSError:
            pass
    return m


def relative_bytes() -> bytes:
    return relative_xlsx(_mtime(hco.IDX_PATH))


def positioning_bytes() -> bytes:
    return positioning_xlsx(_mtime(hco.POS_PATH_FULL, hco.POS_PATH))


def headcount_bytes() -> bytes:
    return headcount_xlsx(_mtime(hco.HC_PATH))


def japan_bytes() -> bytes:
    # cache key = universe yml + snapshots.db mtime（EOD cron 日刷 → 次日自动失效）
    return japan_xlsx(_mtime(hco.JP_UNIVERSE_PATH,
                             hco.REPO_ROOT / "data" / "snapshots.db"))
